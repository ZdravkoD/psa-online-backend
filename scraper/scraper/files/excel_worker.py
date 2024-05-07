import math
import logging
import os
from urllib.parse import urlparse
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from files.file_worker import FileWorker, RowInfo, WorkerProgress
from files.azure_blob_client import AzureBlobClient

# -*- coding: utf-8 -*-

# old format
# NUMBER_OF_ROWS_TO_SKIP_FROM_INPUT_SHEET = 10
# new format
NUMBER_OF_ROWS_TO_SKIP_FROM_INPUT_SHEET = 0
COLUMN_TO_FILL = 0


class ExcelWorker(FileWorker):
    inputFile: openpyxl.Workbook | None = None
    inputSheet: Worksheet | None = None
    inputFilename = ""
    outputFile: openpyxl.Workbook | None = None
    outputSheet: Worksheet | None = None
    outputFilename = ""

    def __init__(self):
        self.currentInputRow = 1 + NUMBER_OF_ROWS_TO_SKIP_FROM_INPUT_SHEET

        self.currentOutputRow = 1
        self.originalProductName = '...'
        self.nrows = 0

        self.notBoughtProducts = []
        self.boughtProducts = []
        # This set is needed in order to ignore duplicate products in our input file
        self.metProducts = set()

    def open_file(self, blob_url: str):
        # Parse the URL to get the path part
        parsed_url = urlparse(blob_url)
        path = parsed_url.path
        # Extract the filename
        self.inputFilename = os.path.basename(path)

        extensionPos = self.inputFilename.find(".xlsx")
        self.outputFilename = self.inputFilename[:extensionPos] + "_Report" + self.inputFilename[extensionPos:]

        fileBytes = AzureBlobClient().download_blob_from_input_container(self.inputFilename)
        with open(self.inputFilename, "wb") as file:
            file.write(fileBytes)
            print("ExcelWorker: File downloaded successfully into: " + self.inputFilename)

        try:
            self.inputFile = openpyxl.load_workbook(self.inputFilename, read_only=True)
            self.inputSheet = self.inputFile.active  # type: ignore
            if self.inputSheet is None:
                raise Exception("ExcelWorker: Can't open input file")
            print("ExcelWorker: Input sheet name: " + self.inputSheet.title)
            print("ExcelWorker: Input file opened successfully from: " + self.inputFilename)
        except Exception as e:
            logging.exception("ExcelWorker: Can't open input file: " + str(e))
            raise Exception("ExcelWorker: Can't open input file: " + str(e))

        self.outputFile = openpyxl.Workbook()
        self.outputSheet = self.outputFile["Sheet"]
        self.outputSheet.title = "Report"

    def _test_writing_to_output_file(self):
        logging.info("ExcelWorker: Test if we can write to output file")
        try:
            if self.outputFile is not None:
                self.outputFile.save(self.outputFilename)
        except Exception as e:
            logging.exception("ExcelWorker: Can't write to output file: " + str(e))
            raise Exception("Моля затворене Report файла и опитайте отново!\n\n" + self.outputFilename)

    def getNumberOfRows(self):
        if self.inputSheet is None:
            raise Exception("ExcelWorker: Input file is not opened")

        return self.inputSheet.max_row

    def get_next_row(self) -> RowInfo:
        if self.inputSheet is None:
            raise Exception("ExcelWorker: Input file is not opened")

        self.nrows = self.getNumberOfRows() + 1
        # self.nrows = 5

        # while self.currentInputRow < self.nrows:
        for row in self.inputSheet.iter_rows(min_row=self.currentInputRow, max_row=self.inputSheet.max_row, max_col=4):

            logging.info("ExcelWorker: Reading row %d from %d", self.currentInputRow, self.nrows)
            # self.markRowInProgress()
            # currentInputRow = self.currentInputRow
            self.currentInputRow += 1

            # self.originalProductName, currentProductNameVariations = self._generateProductNameVariations(
            #     self.inputSheet.cell(row=currentInputRow, column=2).value
            #     )
            print("Product name: " + str(row[1].value), "Product quantity: " + str(row[3].value))
            self.originalProductName, currentProductNameVariations = self._generateProductNameVariations(row[1].value)
            # If the products has been met, ignore it and continue to the next row
            if (self.originalProductName not in self.metProducts):
                try:
                    # value = self.inputSheet.cell(row=currentInputRow, column=4).value
                    value = row[3].value
                    self.currentProductQuantity = int(value)  # type: ignore
                except TypeError:
                    # The quantity of the product is most probably empty. Skip this row
                    self.addNotBoughtProduct(self.originalProductName, -1)
                    continue
                self.metProducts.add(self.originalProductName)
                return RowInfo(self.originalProductName, currentProductNameVariations, self.currentProductQuantity)
            else:
                logging.info("ExcelWorker: Skipping duplicate product: " + self.originalProductName)

        return RowInfo(None, None, None)

    # TODO: Mark progress NOT in the excel file, but in the database
    def markRowInProgress(self):
        if self.inputSheet is None:
            raise Exception("ExcelWorker: Input file is not opened")

        logging.info(f"Marking row: {self.currentInputRow} and column: {COLUMN_TO_FILL + 1}")
        # green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        # self.inputSheet.cell(row=self.currentInputRow, column=COLUMN_TO_FILL + 1).fill = green_fill

    # TODO: Mark progress NOT in the excel file, but in the database
    def unmarkAllRows(self):
        pass
        # for currentRow in range(1, self.nrows + 1):
        #     self._unmarkRow(currentRow)

    # TODO: Mark progress NOT in the excel file, but in the database
    def _unmarkRow(self, row):
        pass
        # if self.inputSheet is None:
        #     raise Exception("ExcelWorker: Input file is not opened")

        # no_fill = PatternFill(start_color='000000', end_color='000000')
        # self.inputSheet.cell(row=row, column=COLUMN_TO_FILL + 1).fill = no_fill

    # TODO: Mark progress NOT in the excel file, but in the database
    def getLastMarkedRow(self):
        pass
        # if self.inputSheet is None:
        #     raise Exception("ExcelWorker: Input file is not opened")

        # last_marked_row = None
        # for row in self.inputSheet.iter_rows():
        #     cell = row[COLUMN_TO_FILL]
        #     if cell.fill and cell.fill.start_color.rgb != '00000000':  # Check if the cell has green fill
        #         last_marked_row = cell.row
        # return last_marked_row

    def _checkValuesForCorrectness(self, currentRow):
        if self.inputSheet is None:
            raise Exception("ExcelWorker: Input file is not opened")

        productName = self.inputSheet.cell(row=currentRow, column=2).value
        if isinstance(productName, str) is not True:
            raise Exception("Проблем с входния Excel файл.\n\nРед: " + str(currentRow)
                            + ", Името на продукта е: " + str(productName) + ". Трябва да бъде валиден текст, а не число."
                            + "\nОбщ брой редове във файла: " + str(self.nrows))

        productQuantity = self.inputSheet.cell(row=currentRow, column=4).value
        if isinstance(productQuantity, float) and productQuantity.is_integer():
            productQuantity = int(productQuantity)
        if isinstance(productQuantity, int) is not True:
            raise Exception("Проблем с входния Excel файл.\n\nРед: " + str(currentRow)
                            + ", Желан брой покупка на продукт е: " + str(productQuantity)
                            + "(" + str(type(productQuantity)) + ")" + ". Трябва да бъде валидно число, а не текст."
                            + "\nОбщ брой редове във файла: " + str(self.nrows))

        if len(str(productName).strip()) == 0:
            raise Exception("Проблем с входния Excel файл.\n\nРед: " + str(currentRow)
                            + ", Името на продукта е: " + str(productName) + ". Полето е празно."
                            + "\nОбщ брой редове във файла: " + str(self.nrows))
        if len(str(productQuantity).strip()) == 0:
            raise Exception("Проблем с входния Excel файл.\n\nРед: " + str(currentRow)
                            + ", Желан брой покупка на продукт е: " + str(productName) + ". Полето е празно."
                            + "\nОбщ брой редове във файла: " + str(self.nrows))

    def validate_input(self):
        self.nrows = self.getNumberOfRows()
        logging.info("Общ брой редове във файла: " + str(self.nrows) + "\n")

        for currentRow in range(1, self.nrows + 1):
            self._checkValuesForCorrectness(currentRow)

    def setProgress(self, newCurrentInputRow):
        self.currentInputRow = newCurrentInputRow

    def get_progress(self) -> WorkerProgress:
        return WorkerProgress(
            self.originalProductName,
            self.currentInputRow - NUMBER_OF_ROWS_TO_SKIP_FROM_INPUT_SHEET - 1,
            self.nrows - NUMBER_OF_ROWS_TO_SKIP_FROM_INPUT_SHEET - 1
        )

    def _writeBoughtProducts(self):
        if self.outputSheet is None:
            raise Exception("ExcelWorker: Output file is not opened")

        fontBold = Font(bold=True)

        # Write header
        self.outputSheet.cell(self.currentOutputRow, 1).value = "Продукт"
        self.outputSheet.cell(self.currentOutputRow, 1).font = fontBold
        self.outputSheet.column_dimensions[get_column_letter(1)].width = 40
        self.outputSheet.cell(self.currentOutputRow, 2).value = "Sting - име на продукт"
        self.outputSheet.cell(self.currentOutputRow, 2).font = fontBold
        self.outputSheet.column_dimensions[get_column_letter(2)].width = 40
        self.outputSheet.cell(self.currentOutputRow, 3).value = "Sting - цена"
        self.outputSheet.cell(self.currentOutputRow, 3).font = fontBold
        self.outputSheet.column_dimensions[get_column_letter(3)].width = 20
        self.outputSheet.cell(self.currentOutputRow, 4).value = "Phoenix - име на продукт"
        self.outputSheet.cell(self.currentOutputRow, 4).font = fontBold
        self.outputSheet.column_dimensions[get_column_letter(4)].width = 40
        self.outputSheet.cell(self.currentOutputRow, 5).value = "Phoenix - цена"
        self.outputSheet.cell(self.currentOutputRow, 5).font = fontBold
        self.outputSheet.column_dimensions[get_column_letter(5)].width = 20
        self.outputSheet.cell(self.currentOutputRow, 6).value = "Добавен в количката на"
        self.outputSheet.cell(self.currentOutputRow, 6).font = fontBold
        self.outputSheet.column_dimensions[get_column_letter(6)].width = 40

        self.currentOutputRow += 1

        # Write content
        for boughProduct in self.boughtProducts:
            self.outputSheet.cell(self.currentOutputRow, 1).value = boughProduct[0]

            self.outputSheet.cell(self.currentOutputRow, 2).value = boughProduct[1]
            self.outputSheet.cell(self.currentOutputRow, 3).value = boughProduct[2]
            self.outputSheet.cell(self.currentOutputRow, 4).value = boughProduct[3]
            self.outputSheet.cell(self.currentOutputRow, 5).value = boughProduct[4]

            if boughProduct[5] == "Sting":
                self.outputSheet.cell(self.currentOutputRow, 2).font = fontBold
                self.outputSheet.cell(self.currentOutputRow, 3).font = fontBold
            if boughProduct[5] == "Phoenix":
                self.outputSheet.cell(self.currentOutputRow, 4).font = fontBold
                self.outputSheet.cell(self.currentOutputRow, 5).font = fontBold

            self.outputSheet.cell(self.currentOutputRow, 6).value = boughProduct[5]
            self.currentOutputRow += 1

    def addNotBoughtProduct(self, product_name: str, product_amount: int):
        self.notBoughtProducts.append((product_name, product_amount))

    def addBoughtProduct(self, original_product_name: str,
                         sting_pharmacy_product_name: str, sting_pharmacy_product_price: float,
                         phoenix_pharmacy_product_name: str, phoenix_pharmacy_product_price: float,
                         bought_from_distributor: str):
        self.boughtProducts.append((original_product_name,
                                    sting_pharmacy_product_name, -1 if sting_pharmacy_product_price == math.inf
                                    else sting_pharmacy_product_price,
                                    phoenix_pharmacy_product_name, -1 if phoenix_pharmacy_product_price == math.inf
                                    else phoenix_pharmacy_product_price,
                                    bought_from_distributor))

    def _writeNotBoughtProducts(self):
        if self.outputFile is None or self.outputSheet is None:
            raise Exception("ExcelWorker: Input file is not opened")

        fontBigBold = Font(bold=True, size=18)
        # Write header
        self.currentOutputRow += 1  # add an empty row
        self.outputSheet.cell(self.currentOutputRow, 1).value = "Списък с некупени продукти"
        self.outputSheet.cell(self.currentOutputRow, 1).font = fontBigBold
        self.currentOutputRow += 1
        # Write content
        for (notBoughProductName, notBoughProductAmount) in self.notBoughtProducts:
            self.outputSheet.cell(self.currentOutputRow, 1).value = notBoughProductName
            self.outputSheet.cell(self.currentOutputRow, 2).value = notBoughProductAmount
            self.currentOutputRow += 1

    def saveOutputFile(self):
        if self.outputFile is None:
            raise Exception("ExcelWorker: Input file is not opened")

        logging.info("ExcelWorker: saveOutputFile")
        self._writeBoughtProducts()
        self._writeNotBoughtProducts()
        self.outputFile.save(self.outputFilename)

    def closeInputFile(self):
        if self.inputFile is None:
            raise Exception("ExcelWorker: Input file is not opened")

        logging.info("ExcelWorker: closeInputFile")
        self.inputFile.save(self.inputFilename)

    def getOutputFilename(self):
        return self.outputFilename
